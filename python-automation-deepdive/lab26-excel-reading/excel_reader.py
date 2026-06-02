import openpyxl

# Task 2: Load Excel file
workbook = openpyxl.load_workbook("sample.xlsx")
sheet = workbook.active

# Task 3: Read specific cell
cell_value = sheet["A1"].value
print(f"The value in cell A1 is: {cell_value}")

print("\n📊 Reading rows:")

for row in sheet.iter_rows(min_row=1, max_col=3, max_row=sheet.max_row, values_only=True):
    print(row)

# Task 4: Sum column B (Score)
sum_column_B = 0

for row in sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row, values_only=True):
    sum_column_B += row[1]

print(f"\n📈 Sum of values in column B (Score): {sum_column_B}")
